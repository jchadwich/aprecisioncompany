import sys
import folium
from math import ceil
import pandas as pd
from pandas import *
from openpyxl import load_workbook
from datetime import datetime
import geopy
from folium.plugins import FloatImage
from geopy.geocoders import Nominatim, ArcGIS, GoogleV3  # Geocoder APIs
import psycopg2

# data from db
conn = psycopg2.connect(
    database="aprecisioncompanydb",
    user="postgres",
    password="infrastructure",
    host="127.0.0.1",
    port="5432",
)
conn.autocommit = True
cursor = conn.cursor()

# Create sorted list of id values to get most recent id for cursor.execute
cursor.execute("""SELECT id from docgen_docgen""")
i = cursor.fetchall()
n = 0  # Index variable
idList = []
for item in i:
    # Pull int values from lists
    idList.append(i[n][0])
    n += 1

cursor.execute(
    f"""SELECT * from docgen_docgen WHERE id = {idList[-1]}"""
)  # Most recent entry only
result = cursor.fetchone()
csvFile = result[7]

today = datetime.now()

# Ensure that the input file is a csv
try:
    userInFile = result[7]
except:
    userInFile = "Delta_Ridge_Townhomes_139926-21Apr2022.csv"

userInFile = pd.read_csv(userInFile)

oidList = userInFile["No."].values.tolist()
lenList = userInFile["Length"].values.tolist()
widList = userInFile["Width"].values.tolist()
scList = userInFile["Special Case"].values.tolist()
qdList = userInFile["Quick Description"].values.tolist()
xList = userInFile["x"].values.tolist()
yList = userInFile["y"].values.tolist()

try:
    saveFile = ""  # result[]
except:
    saveFile = "Delta_Ridge_Townhomes_139926-21Apr2022_proposal.docx"
try:
    inputEntity = result[1]
except:
    inputEntity = "Delta Ridge West and South"
try:
    PPRNum = result[26]
except:
    PPRNum = "11422-211-01"
try:
    PONum = result[25]
except:
    PONum = ""
try:
    indivName = result[10]
except:
    indivName = "Bonnie Giles"
try:
    indivTitle = result[12]
except:
    indivTitle = "Association Manager"
try:
    BDName = result[14]
except:
    BDName = "Jack"
try:
    BDTitle = result[16]
except:
    BDTitle = "BD"
try:
    BDExt = result[13]
except:
    BDExt = "300"
try:
    BDPh = result[15]
except:
    BDPh = "(111)111-000"
try:
    contactAddress = result[8]
except:
    contactAddress = "8305 Falls of Neuse Road, Suite 200, Raleigh, NC 27615"
try:
    contactPh = result[11]
except:
    contactPh = "919-878-8787 ext. 255"
try:
    contactEmail = result[9]
except:
    contactEmail = "bonitagiles@towneproperties.com"
try:
    projName = result[27]
except:
    projName = "Wilmington Proj"
try:
    specs = result[32]
except:
    specs = "½” to 2½”"
try:
    segwaysUsed = result[30]
except:
    segwaysUsed = "no"
try:
    dnrPrice = result[18]
except:
    dnrPrice = 15
try:
    city = result[2]
except:
    city = inputEntity
try:
    safetyIncident = result[29]
except:
    safetyIncident = "0"
try:
    smPrice = result[31]
except:
    smPrice = 2.5
try:
    mdPrice = result[23]
except:
    mdPrice = 5
try:
    lgPrice = result[20]
except:
    lgPrice = 10
try:
    curbPrice = result[17]
except:
    curbPrice = 10
try:
    expNumDaysToRepair = ""  # result[]
except:
    expNumDaysToRepair = 3
try:
    expNumTechs = result[33]
except:
    expNumTechs = 3
try:
    pssMin = result[28]
except:
    pssMin = "$5,000"
try:
    avgPrice = ""  # result[]
except:
    avgPrice = 2.50
try:
    dtcLow = result[24]
except:
    dtcLow = 1
try:
    dtcHigh = result[22]
except:
    dtcHigh = 2
try:
    BDEmail = ""  # result[]
except:
    BDEmail = "bd@precisionsafesidewalks.com"
try:
    workOrderLoc = result[34]
except:
    workOrderLoc = "No"
try:
    mapLayers = result[21]
except:  # ACCEPTABLE INPUT: "dnr", "repairs", "dnr and repair"
    mapLayers = "dnr and repair"

alphaList = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AS",
    "AT",
    "AU",
    "AV",
    "AW",
    "AX",
    "AY",
    "AZ",
    "BA",
    "BB",
    "BC",
    "BD",
    "BE",
    "BF",
    "BG",
    "BH",
    "BI",
    "BJ",
    "BK",
    "BL",
    "BM",
    "BN",
    "BO",
    "BP",
    "BQ",
    "BR",
    "BS",
    "BT",
    "BU",
    "BV",
    "BW",
    "BX",
    "BY",
    "BZ",
]

sqftList = []
costList = []

# Define dicts for unique streets for each deficiency
uniqueStrNames_totH_count = {}
uniqueStrNames_sm_sqft = {}
uniqueStrNames_md_sqft = {}
uniqueStrNames_lg_sqft = {}
uniqueStrNames_curb_length = {}
uniqueStrNames_dnr_sqft = {}
uniqueStrNames_totH_cost = {}
uniqueStrNames_sm_cost = {}
uniqueStrNames_md_cost = {}
uniqueStrNames_lg_cost = {}
uniqueStrNames_curb_cost = {}
uniqueStrNames_dnr_cost = {}
uniqueStrNames_sm_count = {}
uniqueStrNames_md_count = {}
uniqueStrNames_lg_count = {}
uniqueStrNames_curb_count = {}
uniqueStrNames_dnr_count = {}
uniqueStrNames_curb_date = {}
uniqueStrNames_swc_date = {}  # date dictionary for s,m,l repairs
uniqueStrNames_swc_h1 = {}
uniqueStrNames_swc_h2 = {}
uniqueStrNames_swc_sqft = {}
uniqueStrNames_swc_repairedLF = {}
uniqueStrNames_swc_repairedInFt = {}

dateList = []
h1List = []  # Only populate if s,m,l
h2List = []  # Only populate if s,m,l
lengthList_swc = []  # Only populate if s,m,l
widthList_swc = []  # Only populate if s,m,l

centerX = 0
centerY = 0
# Calculate center of map from averages of lats and longs

# Reverse Geocode instantiation
g = ArcGIS()
addresses = []


def Invoice():
    # Declare variables for later calculation
    lgCount = 0
    mdCount = 0
    smCount = 0
    dnrCount = 0
    curbCount = 0
    totalCount = 0
    tripHazardCount = 0

    lgSqft = 0
    mdSqft = 0
    smSqft = 0
    dnrSqft = 0
    totalSqft = 0  # Does not include curb
    curbLength = 0
    tripHazardSqft = 0

    lgCost = 0
    mdCost = 0
    smCost = 0
    dnrCost = 0
    curbCost = 0
    totalCost = 0
    totalDnrCost = 0  # Total cost if all deficiencies were treated as dnr
    savings = 0

    wasteConcreteLow = (
        0  # How many tons were saved from going to the landfill? Low number
    )
    wasteConcreteHigh = (
        0  # How many tons were saved from going to the landfill? High number
    )
    wastConcreteCF = 0  # How many cubic feet of natural resources were saved?
    gasSaved = 0  # How many gallons of gasoline were saved?
    tonsCO2 = 0  # How many tons of CO2 were saved from being pumped into atmosphere?

    for item in qdList:
        if item == "Large":
            lgCount += 1
            tripHazardCount += 1
        elif item == "Medium":
            mdCount += 1
            tripHazardCount += 1
        elif item == "Small":
            smCount += 1
            tripHazardCount += 1
    for item in scList:
        if item == "Replace":
            dnrCount += 1
        elif item == "Curb":
            curbCount += 1
    totalCount = curbCount + dnrCount + smCount + mdCount + lgCount

    # Calculate sqft and length (for curb)
    j = 0
    _fullDEFTYPE_ = []
    for item in qdList:
        if item == "Large":
            _fullDEFTYPE_.append(item)
        elif item == "Medium":
            _fullDEFTYPE_.append(item)
        elif item == "Small":
            _fullDEFTYPE_.append(item)
        else:
            _fullDEFTYPE_.append(scList[j])
        j += 1
    print("full def list")
    print(_fullDEFTYPE_)

    k = 0
    for item in _fullDEFTYPE_:
        if item == "Large":
            lgSqft += float(sqftList[k])
        elif item == "Medium":
            mdSqft += float(sqftList[k])
        elif item == "Small":
            smSqft += float(sqftList[k])
        elif item == "Curb":
            # *******************************
            curbLength += float(
                lenList[k]
            )  ## Is the correct field being used here???????? ****************************!!
            # *******************************
        elif item == "Replace":
            dnrSqft += float(sqftList[k])
        k += 1
        totalSqft = dnrSqft + smSqft + mdSqft + lgSqft
        tripHazardSqft = lgSqft + mdSqft + smSqft

        # Calculate costs

    for item in _fullDEFTYPE_:
        if item == "Large":
            lgCost = lgSqft * lgPrice
        elif item == "Medium":
            mdCost = mdSqft * mdPrice
        elif item == "Small":
            smCost = smSqft * smPrice
        elif item == "Replace":
            dnrCost = dnrSqft * dnrPrice
        elif item == "Curb":
            # ********************************
            curbCost = (
                curbLength * curbPrice
            )  ## Is the correct field being used here???????? ****************************!!
            # ********************************
        totalCost = dnrCost + smCost + mdCost + lgCost + curbCost
        totalHazardCost = smCost + mdCost + lgCost + curbCost

    # Calculate cost if dnr only was used
    totalDnrCost = ((lgSqft + mdSqft + smSqft + dnrSqft) * dnrPrice) + (45 * curbLength)
    savings = (((lgSqft + mdSqft + smSqft) * dnrPrice) + (45 * curbLength)) - (
        smCost + mdCost + lgCost + curbCost
    )

    # Calculate Natural Resource Savings
    tcf = totalSqft * 0.33  # total cubic feet
    tl = (tcf * 132) / 2000  # tons (low)
    th = (tcf * 140) / 2000  # tons (high)

    dump = ceil(th) * 8  # total miles to dump dnr
    bh = ceil(th) * 8  # backhoe transport miles
    totMiles = dump + bh + 20  # total miles
    ga = totMiles / 8  # gallons required to transport to new site
    gan = totMiles / 8  # gallons required to transport to new site
    miscMi = (ceil(dump) / 4) * 6  # miscellaneous mileage
    wf = miscMi / 15  # fuel for workers
    totFuelSaved = wf + ga + gan  # total gallons fossil fuels avoided

    cpCO2 = ((th + tl) / 2) * 0.13078  # CO2 avoided during concrete production
    ffCO2 = totFuelSaved * 0.009  # CO2 avoided fossil fuels
    totCO2 = cpCO2 + ffCO2  # total CO2 avoided

    # Inch ft calculation
    totInFt = 0  # InchFeet = ((H1+H2)/16)*(Length of Panel)*12
    INFT = 0
    # data = read_csv(userInFile)
    # data = pd.read_csv(data)
    data = userInFile
    k = 0
    H1 = data["H1"].tolist()
    H2 = data["H2"].tolist()
    lengthOfPanel = data["Length"].tolist()
    while k < len(H1) - 1:
        if H1[k] != "":
            INFT = ((H1[k] + H2[k]) / 16) * (lengthOfPanel[k])
            k += 1
            totInFt += INFT
    # Other calculations using Inch Feet
    pssConcRem = (totInFt / 50) * 30

    workbook = load_workbook(filename="PS Template - Green 7-26-22 MACRO1.xlsx")
    sheet = workbook.active
    sheet["Q2"] = today
    sheet["P2"] = BDName

    # Curb Repairs Table
    celNum = 20
    for key in uniqueStrNames_curb_date.items():
        celNum_string = "F" + str(celNum)
        sheet[celNum_string] = str(key[0])
        celNum += 1
    celNum = 20
    for value in uniqueStrNames_curb_date.items():
        celNum_string = "E" + str(celNum)
        sheet[celNum_string] = str(value[1])
        celNum += 1
    celNum = 20
    for value in uniqueStrNames_curb_count.items():
        celNum_string = "G" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 20
    for value in uniqueStrNames_curb_length.items():
        celNum_string = "H" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 20
    for value in uniqueStrNames_curb_cost.items():
        celNum_string = "I" + str(celNum)
        sheet[celNum_string] = float(value[1])
        celNum += 1

    # Sidewalk Repairs Table
    celNum = 37
    for key in uniqueStrNames_swc_date.items():
        celNum_string = "B" + str(celNum)
        sheet[celNum_string] = str(key[0])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_swc_date.items():
        celNum_string = "A" + str(celNum)
        sheet[celNum_string] = str(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_totH_count.items():
        celNum_string = "C" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_swc_h1.items():
        celNum_string = "D" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_swc_repairedLF.items():
        celNum_string = "F" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_swc_sqft.items():
        celNum_string = "G" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_swc_repairedInFt.items():
        celNum_string = "H" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_totH_cost.items():
        celNum_string = "I" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1

    # Count and cost by size table
    celNum = 37
    for value in uniqueStrNames_sm_count.items():
        celNum_string = "AH" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_md_count.items():
        celNum_string = "AG" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_lg_count.items():
        celNum_string = "AF" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_sm_cost.items():
        celNum_string = "AK" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_md_cost.items():
        celNum_string = "AJ" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    celNum = 37
    for value in uniqueStrNames_lg_cost.items():
        celNum_string = "AI" + str(celNum)
        sheet[celNum_string] = int(value[1])
        celNum += 1
    ########################################################################################################################
    # Update date sheets
    dateList = []
    # Populate lists:
    # no., H1, H2, length, width, length in inches, tech name, sqft/len(if curb), sidewalk or curb, InFt, cost, date,
    # address location (created already but used here) -- address_loc
    # Only populate if non REPLACE

    Buttons = data[
        "Special Case"
    ].tolist()  # Change this to reflect AGOL field name ??????????????????????
    QD = data[
        "Quick Description"
    ].tolist()  # Change this to reflect AGOL field name ??????????????????????

    lengt = data["Length"].tolist()
    width = data["Width"].tolist()
    sqft = []
    i = 0
    for item in lengt:
        sqft.append(float(item) * float(width[i]))
        i += 1
    h1 = data["H1"].tolist()
    h2 = data["H2"].tolist()
    # oid = data['OBJECTID'].tolist() # If necessary, change unique identifier field name to "OBJECTID"S
    oid = data[
        "No."
    ].tolist()  # If necessary, change unique identifier field name to "OBJECTID"S
    date = data[
        "Date"
    ].tolist()  # Change formatting of cells to include date only, not time
    len_in_inches = []
    for item in lengt:
        len_in_inches.append(int(item) * 12)
    # tech_name = data['Creator'].tolist()
    # Temp for testing, change back to creator at rollout
    tech_name = "RR"
    inft = []
    i = 0
    for item in h1:
        inft.append(((float(item) + h2[i]) / 16) * lengt[i])
        i += 1
    sw_c = []
    i = 0
    while i < len(QD):
        if QD[i] == "Small":
            sw_c.append(QD[i])
            i += 1
        elif QD[i] == "Medium":
            sw_c.append(QD[i])
            i += 1
        elif QD[i] == "Small":
            sw_c.append(QD[i])
            i += 1
        else:
            sw_c.append(Buttons[i])
            i += 1
    _cost_ = []
    i = 0
    for item in sw_c:
        if item == "Small":
            _cost_.append(lengt[i] * width[i] * smPrice)
        elif item == "Medium":
            _cost_.append(lengt[i] * width[i] * mdPrice)
        elif item == "Large":
            _cost_.append(lengt[i] * width[i] * lgPrice)
        elif item == "Curb":
            _cost_.append(lengt[i] * width[i] * curbPrice)
        else:
            _cost_.append("")
    # Create list of road names AND NUMBERS from addresses list
    itemList = []
    newLineList = []
    address_loc = []
    for item in addresses:
        itemList.append(str(item))
    for item in itemList:
        strippedItem = item.strip()
        ni = strippedItem.split(",")
        address_loc.append(ni[0])

    ########################################################################################################################
    # get a list of unique tech names
    unique_tech_names = []
    for item in tech_name:
        if item not in unique_tech_names:
            unique_tech_names.append((item))
    # get list of unique dates
    date_name = []
    for item in date:
        if item not in date_name:
            date_name.append(item)

    # Create dictionaries
    date_length = {}
    for item in date:
        if item not in date_length:
            date_length[item] = []
    date_width = {}
    for item in date_name:
        if item not in date_width:
            date_width[item] = []
    date_oid = {}
    for item in date_name:
        if item not in date_oid:
            date_oid[item] = []
    date_h1 = {}
    for item in date_name:
        if item not in date_h1:
            date_h1[item] = []
    date_h2 = {}
    for item in date_name:
        if item not in date_h2:
            date_h2[item] = []
    date_techN = {}
    for item in date_name:
        if item not in date_techN:
            date_techN[item] = []
    date_address_loc = {}
    for item in date_name:
        if item not in date_address_loc:
            date_address_loc[item] = []

    # Fill dictionaries with appropriate values
    i = 0
    for item in date:
        date_length[item].append(lengt[i])
        i += 1
    i = 0
    for item in date:
        date_width[item].append(width[i])
        i += 1
    i = 0
    for item in date:
        date_oid[item].append(oid[i])
        i += 1
    i = 0
    for item in date:
        date_h1[item].append(h1[i])
        i += 1
    i = 0
    for item in date:
        date_h2[item].append(h2[i])
        i += 1
    i = 0
    for item in date:
        date_techN[item].append(tech_name[i])
        i += 1
    i = 0
    for item in date:
        date_address_loc[item].append(address_loc[i])
        i += 1

    # Add the above information to the appropriate sheets here

    alphaIndex = 0
    for key in date_length.items():
        ws = workbook[alphaList[alphaIndex].lower()]

        listLoc = 25
        rowLoc = 4
        for item in unique_tech_names:
            celLoc = alphaList[listLoc] + str(rowLoc)
            print(celLoc)
            ws[celLoc] = str(item)
            listLoc += 1

        i = 0
        row = 22
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "B" + str(row)
            ws[cn] = int(item)
            i += 1
            row += 1
        alphaIndex += 1
    alphaIndex = 0
    for key in date_width.items():
        ws = workbook[alphaList[alphaIndex].lower()]
        i = 0
        row = 22
        ws["A1"] = key[0]
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "A" + str(row)
            ws[cn] = int(item)
            i += 1
            row += 1
        alphaIndex += 1
    alphaIndex = 0
    for key in date_h1.items():
        ws = workbook[alphaList[alphaIndex].lower()]
        i = 0
        row = 22
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "D" + str(row)
            ws[cn] = int(item)
            i += 1
            row += 1
        alphaIndex += 1
    alphaIndex = 0
    for key in date_h2.items():
        ws = workbook[alphaList[alphaIndex].lower()]
        i = 0
        row = 22
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "E" + str(row)
            ws[cn] = int(item)
            i += 1
            row += 1
        alphaIndex += 1
    alphaIndex = 0
    for key in date_address_loc.items():
        ws = workbook[alphaList[alphaIndex].lower()]
        i = 0
        row = 22
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "G" + str(row)
            ws[cn] = item
            i += 1
            row += 1
        alphaIndex += 1
    alphaIndex = 0
    for key in date_techN.items():
        ws = workbook[alphaList[alphaIndex].lower()]
        i = 0
        row = 22
        newList = key[1]  # create list outside dictionary
        for item in newList:
            cn = "L" + str(row)
            ws[cn] = item
            i += 1

            row += 1
        alphaIndex += 1

    workbook.save(filename=inputEntity + "_Invoice.xlsx")


m = Map()
i = Invoice()
